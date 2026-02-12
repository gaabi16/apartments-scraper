import time
import random
import re
import os
import sys
import tempfile
from openpyxl import Workbook
from playwright.sync_api import sync_playwright

# Import Database
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import Database.database as database

def clean_text(text):
    if not text: return None
    cleaned = " ".join(text.replace("\n", " ").split())
    return cleaned if cleaned else None

def extract_price(text):
    if not text: return None
    # Păstrăm doar cifrele
    digits = re.sub(r'[^\d]', '', text)
    if not digits: return None
    return int(digits)

def extract_surface(text):
    if not text: return None
    # Căutăm pattern: 50 mp, 50.5 m2 etc.
    match = re.search(r'(\d+(?:[.,]\d+)?)\s*(?:mp|m2|metri)', text, re.IGNORECASE)
    if match:
        val_str = match.group(1).replace(',', '.')
        try:
            return float(val_str)
        except:
            return None
    return None

def handle_cookie_consent(page):
    """Încearcă să închidă bannerul de cookie."""
    try:
        # Așteptăm selectorul specific OneTrust sau text generic
        page.wait_for_selector("button#onetrust-accept-btn-handler, button:has-text('Accept'), button:has-text('De acord')", timeout=3000)
        # Click pe primul găsit
        if page.is_visible("button#onetrust-accept-btn-handler"):
            page.click("button#onetrust-accept-btn-handler")
        else:
            page.click("button:has-text('Accept')")
        print("Banner cookie închis.")
    except:
        print("Bannerul cookie nu a apărut sau a fost deja închis.")

def scrape_detail_page(context, url):
    """
    Deschide o pagină nouă (tab) pentru detaliile anunțului,
    extrage datele și o închide.
    """
    data = {
        "description": None,
        "floor": None,
        "contact_name": None,
        "phone_number": None,
        "rooms": None
    }
    
    page = context.new_page()
    try:
        page.goto(url, wait_until="domcontentloaded")
        # Pauză mică pentru încărcare elemente dinamice
        page.wait_for_timeout(random.randint(1500, 3000))

        # --- 1. Extragere Telefon ---
        try:
            # Selector pentru butonul de telefon
            phone_btn_sel = "button.detail-listing-open-phone-modal"
            if page.is_visible(phone_btn_sel):
                # Click pentru a dezvălui numărul
                page.click(phone_btn_sel)
                page.wait_for_timeout(1000) # Așteptăm să apară textul
                
                # Extragem textul din buton sau span-ul interior
                phone_text = page.locator(phone_btn_sel).inner_text()
                # Curățăm textul pentru a găsi numărul
                extracted = re.search(r'(\d{10}|\d{3}\s\d{3}\s\d{3})', phone_text)
                if extracted:
                    data["phone_number"] = extracted.group(0).replace(" ", "")
        except Exception as e:
            # print(f"Nu s-a putut extrage telefonul: {e}")
            pass

        # --- 2. Descriere ---
        try:
            desc_loc = page.locator(".text-content").first
            if not desc_loc.is_visible():
                desc_loc = page.locator("#truncatedDescription").first
            
            if desc_loc.is_visible():
                data["description"] = clean_text(desc_loc.inner_text())
        except:
            pass

        # --- 3. Detalii (Etaj, Camere) ---
        # Iterăm prin elementele de specificații
        try:
            specs = page.locator("span.text-grey-700").all()
            for spec in specs:
                txt = spec.inner_text().strip()
                # Mergem la părinte pentru a găsi valoarea (care e un sibling sau child al părintelui)
                # Structura uzuală: <div><span>Etaj:</span> <span class="font-bold">1</span></div>
                # Playwright permite selectarea părintelui prin xpath sau locator chaining
                parent = spec.locator("..") 
                val_loc = parent.locator(".font-semibold, .font-bold").first
                
                if val_loc.is_visible():
                    val_text = clean_text(val_loc.inner_text())
                    
                    if "Etaj:" in txt:
                        data["floor"] = val_text
                    elif "Nr. cam.:" in txt:
                        try:
                            data["rooms"] = int(val_text)
                        except:
                            pass
        except:
            pass

        # --- 4. Contact ---
        try:
            # Nume agent/proprietar
            contact_loc = page.locator("p.text-base.font-bold").first
            if contact_loc.is_visible():
                data["contact_name"] = clean_text(contact_loc.inner_text())
            else:
                agency_loc = page.locator("[data-cy='agency-name']").first
                if agency_loc.is_visible():
                    data["contact_name"] = clean_text(agency_loc.inner_text())
        except:
            pass

    except Exception as e:
        print(f"Eroare la pagina de detaliu {url}: {e}")
    finally:
        page.close()

    return data

def scrape_imobiliare(rooms, price_min, price_max, sector):
    # Setup path Excel
    tmp = tempfile.gettempdir()
    file_path = os.path.join(tmp, f"imobiliare_s{sector}_{int(time.time())}.xlsx")
    results_to_save = []

    # Configurare Playwright
    with sync_playwright() as p:
        # Lansăm browserul (headless=False ca să vezi ce face, poți pune True pentru producție)
        browser = p.chromium.launch(headless=False, args=["--start-maximized"])
        
        # Context cu User Agent de om real
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 720}
        )
        
        page = context.new_page()

        try:
            room_str = f"{rooms}-camere" if rooms > 1 else "1-camera"
            base_url = f"https://www.imobiliare.ro/vanzare-apartamente/bucuresti/sector-{sector}/{room_str}"
            params = f"?price={price_min}-{price_max}&floor=1%2C2%2C3%2C4%2C5%2C6%2C7%2C8%2C9%2C10%2Cabove-10%2Cexcluded-last-floor"
            full_url = base_url + params
            
            print(f"1. Accesare URL: {full_url}")
            page.goto(full_url, timeout=60000)
            
            # Verificare Anti-Bot
            if "anomalie" in page.content().lower() or "captcha" in page.content().lower():
                print("!!! BLOCAJ DETECTAT !!! - Schimbă IP-ul.")
                wb = Workbook()
                ws = wb.active
                ws.append(["IP Blocat. Schimbă rețeaua și reîncearcă."])
                wb.save(file_path)
                return file_path

            handle_cookie_consent(page)

            # Scroll pentru încărcare
            print("2. Derulare pagină pentru încărcare anunțuri...")
            for _ in range(4): # Scroll de câteva ori
                page.mouse.wheel(0, 3000)
                page.wait_for_timeout(1500)
            
            # Colectare link-uri din listă
            print("3. Colectare link-uri...")
            unique_candidates = {}
            
            # Selectăm toate cardurile
            cards = page.locator('div[id^="listing-"]').all()
            print(f"   -> S-au găsit vizual {len(cards)} carduri.")

            for card in cards:
                try:
                    # Link
                    link_loc = card.locator('a[data-cy="listing-information-link"]').first
                    if not link_loc.is_visible():
                        link_loc = card.locator('h2 a').first
                    
                    if not link_loc.is_visible(): continue
                    
                    link = link_loc.get_attribute("href")
                    if not link.startswith("http"):
                        link = "https://www.imobiliare.ro" + link

                    # Preț
                    price_loc = card.locator(".price, [data-cy='card-price']").first
                    price_val = 0
                    if price_loc.is_visible():
                        price_val = extract_price(price_loc.inner_text())

                    # Filtrare preliminară
                    if not price_val or not (price_min <= price_val <= price_max):
                        continue

                    # Titlu
                    title_loc = card.locator("span.relative, h2").first
                    title = clean_text(title_loc.inner_text()) if title_loc.is_visible() else "Titlu"

                    # Locație
                    loc_loc = card.locator(".location, p.capitalize").first
                    location = clean_text(loc_loc.inner_text()) if loc_loc.is_visible() else ""

                    # Suprafață
                    full_text = card.inner_text()
                    surface_val = extract_surface(full_text)

                    fingerprint = f"{title}_{location}_{price_val}_{surface_val}"
                    
                    if fingerprint not in unique_candidates:
                        unique_candidates[fingerprint] = {
                            "title": title, "location": location, 
                            "price": price_val, "surface": surface_val, 
                            "link": link, "rooms_initial": rooms
                        }
                except:
                    continue

            print(f"4. Începe extragerea detaliată pentru {len(unique_candidates)} anunțuri...")

            # Vizitare fiecare anunț
            items = list(unique_candidates.values())
            
            for idx, item in enumerate(items):
                print(f"   [{idx+1}/{len(items)}] Procesare: {item['link']}")
                
                details = scrape_detail_page(context, item['link'])
                
                final_obj = {
                    'source_website': 'Imobiliare.ro',
                    'title': item['title'],
                    'price': item['price'],
                    'location': item['location'],
                    'surface': item['surface'],
                    'link': item['link'],
                    'description': details['description'] if details['description'] else item['title'],
                    'floor': details['floor'],
                    'contact_name': details['contact_name'],
                    'phone_number': details['phone_number'],
                    'rooms': details['rooms'] if details['rooms'] else item['rooms_initial']
                }
                results_to_save.append(final_obj)

        except Exception as e:
            print(f"Eroare generală Playwright: {e}")
        finally:
            browser.close()

    # Salvare în DB și Excel
    if results_to_save:
        print(f"5. Se salvează {len(results_to_save)} rezultate în DB...")
        try:
            database.insert_batch_apartments(results_to_save)
        except Exception as e:
            print(f"Eroare DB: {e}")

        wb = Workbook()
        ws = wb.active
        ws.append(["Titlu", "Descriere", "Pret", "Locatie", "Suprafata", "Etaj", "Camere", "Nume Contact", "Telefon", "Link"])
        
        for r in results_to_save:
            ws.append([
                r.get('title'), r.get('description'), r.get('price'), 
                r.get('location'), r.get('surface'), r.get('floor'), 
                r.get('rooms'), r.get('contact_name'), r.get('phone_number'), 
                r.get('link')
            ])
        
        wb.save(file_path)
        print(f"6. Excel generat: {file_path}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nu au fost gasite rezultate (verificati filtrele sau IP-ul)"])
        wb.save(file_path)
        print("Nu au fost găsite rezultate.")

    return file_path